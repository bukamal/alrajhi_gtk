# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTextEdit,
                             QComboBox, QLabel, QDialog, QFormLayout, QTabWidget, QFrame,
                             QDateEdit, QProgressBar, QMessageBox, QFileDialog, QApplication)
from PyQt5.QtCore import Qt, QDate
from database import db
from utils_pyqt5 import format_currency, show_toast
import os

class ReportsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(8,8,8,8)

        # شريط الفلاتر والأزرار
        filter_bar = QFrame()
        filter_bar.setStyleSheet("QFrame { background-color: #f8fafc; border-radius: 12px; padding: 8px; }")
        filter_layout = QHBoxLayout(filter_bar)
        filter_layout.setContentsMargins(12, 8, 12, 8)

        filter_layout.addWidget(QLabel("من تاريخ:"))
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-90))
        self.start_date.setCalendarPopup(True)
        filter_layout.addWidget(self.start_date)

        filter_layout.addWidget(QLabel("إلى تاريخ:"))
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        filter_layout.addWidget(self.end_date)

        self.apply_filter_btn = QPushButton("تطبيق الفلترة")
        self.apply_filter_btn.clicked.connect(self.refresh_all)
        filter_layout.addWidget(self.apply_filter_btn)

        self.refresh_btn = QPushButton("🔄 تحديث")
        self.refresh_btn.clicked.connect(self.refresh_all)
        filter_layout.addWidget(self.refresh_btn)

        self.export_pdf_btn = QPushButton("📄 تصدير PDF (التقرير الحالي)")
        self.export_pdf_btn.clicked.connect(self.export_current_to_pdf)
        filter_layout.addWidget(self.export_pdf_btn)

        self.export_excel_btn = QPushButton("📊 تصدير Excel (التقرير الحالي)")
        self.export_excel_btn.clicked.connect(self.export_current_to_excel)
        filter_layout.addWidget(self.export_excel_btn)

        self.print_btn = QPushButton("🖨️ طباعة")
        self.print_btn.clicked.connect(self.print_current_report)
        filter_layout.addWidget(self.print_btn)

        filter_layout.addStretch()
        layout.addWidget(filter_bar)

        # شريط تقدم (مخفي عادة)
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setDocumentMode(True)
        self.tabs.currentChanged.connect(self.on_tab_changed)

        # تبويب ميزان المراجعة
        self.trial_tab = QWidget()
        self.trial_layout = QVBoxLayout(self.trial_tab)
        self.trial_text = QTextEdit()
        self.trial_text.setReadOnly(True)
        self.trial_layout.addWidget(self.trial_text)
        self.tabs.addTab(self.trial_tab, "📊 ميزان المراجعة")

        # تبويب قائمة الدخل
        self.income_tab = QWidget()
        self.income_layout = QVBoxLayout(self.income_tab)
        self.income_text = QTextEdit()
        self.income_text.setReadOnly(True)
        self.income_layout.addWidget(self.income_text)
        self.tabs.addTab(self.income_tab, "💰 قائمة الدخل")

        # تبويب الميزانية العمومية
        self.balance_tab = QWidget()
        self.balance_layout = QVBoxLayout(self.balance_tab)
        self.balance_text = QTextEdit()
        self.balance_text.setReadOnly(True)
        self.balance_layout.addWidget(self.balance_text)
        self.tabs.addTab(self.balance_tab, "🏦 الميزانية العمومية")

        # تبويب كشف حساب عميل
        self.customer_tab = QWidget()
        self.customer_layout = QVBoxLayout(self.customer_tab)
        customer_top = QHBoxLayout()
        self.customer_combo = QComboBox()
        self.customer_combo.setMinimumWidth(250)
        customer_top.addWidget(QLabel("اختر العميل:"))
        customer_top.addWidget(self.customer_combo)
        customer_top.addStretch()
        self.customer_layout.addLayout(customer_top)
        self.customer_text = QTextEdit()
        self.customer_text.setReadOnly(True)
        self.customer_layout.addWidget(self.customer_text)
        self.tabs.addTab(self.customer_tab, "👤 كشف حساب عميل")

        # تبويب كشف حساب مورد
        self.supplier_tab = QWidget()
        self.supplier_layout = QVBoxLayout(self.supplier_tab)
        supplier_top = QHBoxLayout()
        self.supplier_combo = QComboBox()
        self.supplier_combo.setMinimumWidth(250)
        supplier_top.addWidget(QLabel("اختر المورد:"))
        supplier_top.addWidget(self.supplier_combo)
        supplier_top.addStretch()
        self.supplier_layout.addLayout(supplier_top)
        self.supplier_text = QTextEdit()
        self.supplier_text.setReadOnly(True)
        self.supplier_layout.addWidget(self.supplier_text)
        self.tabs.addTab(self.supplier_tab, "🏭 كشف حساب مورد")

        layout.addWidget(self.tabs)

        self.refresh_customers()
        self.refresh_suppliers()
        self.refresh_all()

    def refresh_all(self):
        self.show_progress(True)
        self.show_trial_balance()
        self.show_income_statement()
        self.show_balance_sheet()
        self.refresh_customers()
        self.refresh_suppliers()
        self.show_progress(False)

    def show_progress(self, show):
        self.progress.setVisible(show)
        if show:
            self.progress.setRange(0, 0)  # غير محدد
        QApplication.processEvents()

    def on_tab_changed(self, index):
        pass

    def refresh_customers(self):
        customers = db.get_customers()
        self.customer_combo.clear()
        for c in customers:
            self.customer_combo.addItem(f"{c['name']} (الرصيد: {format_currency(c['balance'])})", c['id'])
        if customers:
            self.show_customer_statement()

    def refresh_suppliers(self):
        suppliers = db.get_suppliers()
        self.supplier_combo.clear()
        for s in suppliers:
            self.supplier_combo.addItem(f"{s['name']} (الرصيد: {format_currency(s['balance'])})", s['id'])
        if suppliers:
            self.show_supplier_statement()

    def get_date_range(self):
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")
        return start, end

    def show_trial_balance(self):
        start, end = self.get_date_range()
        trial = db.get_trial_balance_filtered(start, end)
        html = self._trial_to_html(trial)
        self.trial_text.setHtml(html)

    def show_income_statement(self):
        start, end = self.get_date_range()
        stmt = db.get_income_statement_filtered(start, end)
        html = self._income_to_html(stmt)
        self.income_text.setHtml(html)

    def show_balance_sheet(self):
        start, end = self.get_date_range()
        bs = db.get_balance_sheet_filtered(start, end)
        html = self._balance_to_html(bs)
        self.balance_text.setHtml(html)

    def show_customer_statement(self):
        cust_id = self.customer_combo.currentData()
        if not cust_id:
            self.customer_text.setHtml("<div style='text-align: center;'><h3>اختر عميلاً أولاً</h3></div>")
            return
        lines = db.get_customer_statement(cust_id)
        html = self._statement_to_html(lines, self.customer_combo.currentText().split(' (')[0], "عميل")
        self.customer_text.setHtml(html)

    def show_supplier_statement(self):
        supp_id = self.supplier_combo.currentData()
        if not supp_id:
            self.supplier_text.setHtml("<div style='text-align: center;'><h3>اختر مورداً أولاً</h3></div>")
            return
        lines = db.get_supplier_statement(supp_id)
        html = self._statement_to_html(lines, self.supplier_combo.currentText().split(' (')[0], "مورد")
        self.supplier_text.setHtml(html)

    def _trial_to_html(self, trial):
        html = """
        <div style="text-align: center;">
            <h2 style="color: #2c3e50;">ميزان المراجعة</h2>
            <hr>
        </div>
        <table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
            <thead>
                <tr style="background-color: #34495e; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">مدين</th>
                    <th style="padding: 8px; text-align: center;">دائن</th>
                </tr>
            </thead>
            <tbody>
        """
        for row in trial:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px; text-align: center;">{row['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(row['debit'])}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(row['credit'])}浏
                </tr>
            """
        total_debit = sum(r['debit'] for r in trial)
        total_credit = sum(r['credit'] for r in trial)
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px; text-align: center;">الإجمالي浏
                    <td style="padding: 8px; text-align: center;">{format_currency(total_debit)}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(total_credit)}浏
                </tr>
            </tbody>
        </table>
        """
        return html

    def _income_to_html(self, stmt):
        html = """
        <div style="text-align: center;">
            <h2 style="color: #2c3e50;">قائمة الدخل</h2>
            <hr>
        </div>
        <h3 style="color: #27ae60;">الإيرادات</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
            <thead>
                <tr style="background-color: #2ecc71; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for inc in stmt['income']:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px;">{inc['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(inc['balance'])}浏
                </tr>
            """
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px;">إجمالي الإيرادات浏
                    <td style="padding: 8px; text-align: center;">{format_currency(stmt['total_income'])}浏
                </tr>
            </tbody>
        </table>
        <h3 style="color: #e67e22;">المصروفات</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
            <thead>
                <tr style="background-color: #f39c12; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for exp in stmt['expenses']:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px;">{exp['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(exp['balance'])}浏
                </tr>
            """
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px;">إجمالي المصروفات浏
                    <td style="padding: 8px; text-align: center;">{format_currency(stmt['total_expenses'])}浏
                </tr>
                <tr style="background-color: #3498db; color: white; font-weight: bold;">
                    <td style="padding: 8px;">صافي الربح浏
                    <td style="padding: 8px; text-align: center;">{format_currency(stmt['net_profit'])}浏
                </tr>
            </tbody>
        </table>
        """
        return html

    def _balance_to_html(self, bs):
        html = """
        <div style="text-align: center;">
            <h2 style="color: #2c3e50;">الميزانية العمومية</h2>
            <hr>
        </div>
        <h3 style="color: #2980b9;">الأصول</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
            <thead>
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for a in bs['assets']:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px;">{a['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(a['debit'])}浏
                </tr>
            """
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px;">إجمالي الأصول浏
                    <td style="padding: 8px; text-align: center;">{format_currency(bs['total_assets'])}浏
                </tr>
            </tbody>
        </table>
        <h3 style="color: #e67e22;">الخصوم</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
            <thead>
                <tr style="background-color: #f39c12; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for l in bs['liabilities']:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px;">{l['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(l['credit'])}浏
                </tr>
            """
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px;">إجمالي الخصوم浏
                    <td style="padding: 8px; text-align: center;">{format_currency(bs['total_liabilities'])}浏
                </tr>
            </tbody>
        </table>
        <h3 style="color: #27ae60;">حقوق الملكية</h3>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
            <thead>
                <tr style="background-color: #2ecc71; color: white;">
                    <th style="padding: 8px; text-align: center;">الحساب</th>
                    <th style="padding: 8px; text-align: center;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for e in bs['equity']:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px;">{e['name']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(e['credit'])}浏
                </tr>
            """
        html += f"""
                <tr style="background-color: #ecf0f1; font-weight: bold;">
                    <td style="padding: 8px;">إجمالي حقوق الملكية浏
                    <td style="padding: 8px; text-align: center;">{format_currency(bs['total_equity'])}浏
                </tr>
            </tbody>
        </table>
        """
        return html

    def _statement_to_html(self, lines, entity_name, entity_type):
        if not lines:
            return f"<div style='text-align: center;'><h3>لا توجد حركات لهذا {entity_type}</h3></div>"
        html = f"""
        <div style="text-align: center;">
            <h2 style="color: #2c3e50;">كشف حساب {entity_type}</h2>
            <h3>{entity_name}</h3>
            <hr>
        </div>
        <table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
            <thead>
                <tr style="background-color: #34495e; color: white;">
                    <th style="padding: 8px;">التاريخ</th>
                    <th style="padding: 8px;">الوصف</th>
                    <th style="padding: 8px;">مدين</th>
                    <th style="padding: 8px;">دائن</th>
                    <th style="padding: 8px;">الرصيد</th>
                </tr>
            </thead>
            <tbody>
        """
        for l in lines:
            html += f"""
                <tr style="border-bottom: 1px solid #ddd;">
                    <td style="padding: 8px; text-align: center;">{l['date']}浏
                    <td style="padding: 8px;">{l['description']}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(l['debit'])}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(l['credit'])}浏
                    <td style="padding: 8px; text-align: center;">{format_currency(l['balance'])}浏
                </tr>
            """
        html += """
            </tbody>
        </table>
        """
        return html

    def get_current_html(self):
        current = self.tabs.currentWidget()
        if current == self.trial_tab:
            return self.trial_text.toHtml()
        elif current == self.income_tab:
            return self.income_text.toHtml()
        elif current == self.balance_tab:
            return self.balance_text.toHtml()
        elif current == self.customer_tab:
            return self.customer_text.toHtml()
        elif current == self.supplier_tab:
            return self.supplier_text.toHtml()
        else:
            return ""

    def export_current_to_pdf(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للتصدير", "error", self)
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير كـ PDF", "report.pdf", "PDF (*.pdf)")
        if not filename:
            return
        from PyQt5.QtGui import QTextDocument
        from PyQt5.QtPrintSupport import QPrinter
        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(filename)
        doc.print(printer)
        show_toast("تم حفظ التقرير كـ PDF", "success", self)

    def export_current_to_excel(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للتصدير", "error", self)
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير كـ Excel", "report.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            show_toast("مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl", "error", self)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير"
        ws.cell(row=1, column=1, value="تم تصدير التقرير من نظام الراجحي")
        ws.cell(row=2, column=1, value="يرجى نسخ البيانات يدوياً من التقرير الأصلي")
        wb.save(filename)
        show_toast("تم حفظ التقرير كـ Excel (مبسط)", "success", self)

    def print_current_report(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للطباعة", "error", self)
            return
        from PyQt5.QtGui import QTextDocument
        from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(lambda p: doc.print(p))
        preview.exec()
