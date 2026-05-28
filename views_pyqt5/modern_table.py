# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import QTableView, QHeaderView, QMenu, QAction, QFileDialog, QMessageBox, QApplication
from PyQt5.QtCore import Qt, QModelIndex, QSortFilterProxyModel, pyqtSignal
from PyQt5.QtGui import QKeySequence
import sys

class ModernTableView(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.setSelectionMode(QTableView.ExtendedSelection)
        self.setSortingEnabled(True)
        self.setWordWrap(True)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.horizontalHeader().setStretchLastSection(True)
        self.verticalHeader().setVisible(False)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.setStyleSheet("""
            QTableView {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                selection-background-color: #dbeafe;
                alternate-background-color: #f8fafc;
            }
            QHeaderView::section {
                background-color: #f1f5f9;
                font-weight: bold;
                border-bottom: 2px solid #3b82f6;
                padding: 8px;
            }
        """)
        self.copy_action = QAction(self)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.copy_action.triggered.connect(self.copy_selection)
        self.addAction(self.copy_action)

    def _show_context_menu(self, pos):
        menu = QMenu()
        export_excel = QAction("📊 تصدير إلى Excel", self)
        export_excel.triggered.connect(self.export_to_excel)
        menu.addAction(export_excel)
        export_pdf = QAction("📄 طباعة", self)
        export_pdf.triggered.connect(self.print_table)
        menu.addAction(export_pdf)
        menu.addSeparator()
        copy_action = QAction("📋 نسخ", self)
        copy_action.triggered.connect(self.copy_selection)
        menu.addAction(copy_action)
        menu.exec(self.viewport().mapToGlobal(pos))

    def copy_selection(self):
        selection = self.selectionModel().selectedIndexes()
        if not selection:
            return
        rows = sorted(set(idx.row() for idx in selection))
        cols = sorted(set(idx.column() for idx in selection))
        text = ""
        for row in rows:
            row_data = []
            for col in cols:
                idx = self.model().index(row, col)
                data = self.model().data(idx, Qt.DisplayRole)
                row_data.append(str(data))
            text += "\t".join(row_data) + "\n"
        QApplication.clipboard().setText(text)

    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(self, "تحذير", "مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl")
            return
        model = self.model()
        if not model:
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير", "report.xlsx", "Excel Files (*.xlsx)")
        if not filename:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير"
        headers = [model.headerData(i, Qt.Horizontal, Qt.DisplayRole) for i in range(model.columnCount())]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row in range(model.rowCount()):
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                value = model.data(idx, Qt.DisplayRole)
                ws.cell(row=row+2, column=col+1, value=value)
        wb.save(filename)
        QMessageBox.information(self, "نجاح", f"تم تصدير البيانات إلى {filename}")

    def print_table(self):
        from PyQt5.QtGui import QTextDocument, QTextCursor
        from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
        model = self.model()
        if not model:
            return
        html = "<html><head><meta charset='UTF-8'></head><body><table border='1' cellpadding='5' style='border-collapse:collapse;'>"
        html += "<thead><tr>"
        for col in range(model.columnCount()):
            header = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
            html += f"<th>{header}</th>"
        html += "</thead><tbody>"
        for row in range(model.rowCount()):
            html += "<tr>"
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                value = model.data(idx, Qt.DisplayRole)
                html += f"<td>{value}</td>"
            html += "</tr>"
        html += "</tbody><td></body></html>"
        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(lambda p: doc.print(p))
        preview.exec()

class SortFilterProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._filter_text = ""

    def setFilterText(self, text):
        self._filter_text = text.lower()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        if not self._filter_text:
            return True
        model = self.sourceModel()
        for col in range(model.columnCount()):
            idx = model.index(source_row, col)
            data = model.data(idx, Qt.DisplayRole)
            if data and self._filter_text in str(data).lower():
                return True
        return False
