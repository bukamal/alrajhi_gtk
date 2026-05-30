# views_pyqt5/reports_widget.py
# -*- coding: utf-8 -*-

import os
import sys
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTextEdit,
                             QComboBox, QLabel, QTabWidget, QFrame, QDateEdit, QProgressBar,
                             QFileDialog, QApplication, QMessageBox)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QTextDocument
from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
from database import reporting_dao, customer_dao, supplier_dao
from utils_pyqt5 import format_currency, show_toast
from views_pyqt5.centered_dialog import show_centered_messagebox
from views_pyqt5.reports import (TrialBalanceReport, IncomeStatementReport,
                                 BalanceSheetReport, CustomerStatementReport,
                                 SupplierStatementReport)

class ReportsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(8,8,8,8)

        filter_bar = QFrame()
        filter_bar.setStyleSheet("QFrame { background-color: #f8fafc; border-radius: 12px; padding: 8px; }")
        filter_layout = QHBoxLayout(filter_bar)
        filter_layout.setContentsMargins(12, 8, 12, 8)

        filter_layout.addWidget(QLabel("من تاريخ:"))
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-90))
        self.start_date.setCalendarPopup(True)
        filter_layout.addWidget(self.start_date)

        filter_layout.addWidget(QLabel("إلى تاريخ:"))
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        filter_layout.addWidget(self.end_date)

        self.apply_filter_btn = QPushButton("تطبيق الفلترة")
        self.apply_filter_btn.clicked.connect(self.refresh_all)
        filter_layout.addWidget(self.apply_filter_btn)

        self.refresh_btn = QPushButton("🔄 تحديث")
        self.refresh_btn.clicked.connect(self.refresh_all)
        filter_layout.addWidget(self.refresh_btn)

        self.export_pdf_btn = QPushButton("📄 تصدير PDF (التقرير الحالي)")
        self.export_pdf_btn.clicked.connect(self.export_current_to_pdf)
        filter_layout.addWidget(self.export_pdf_btn)

        self.export_excel_btn = QPushButton("📊 تصدير Excel (التقرير الحالي)")
        self.export_excel_btn.clicked.connect(self.export_current_to_excel)
        filter_layout.addWidget(self.export_excel_btn)

        self.print_btn = QPushButton("🖨️ طباعة")
        self.print_btn.clicked.connect(self.print_current_report)
        filter_layout.addWidget(self.print_btn)

        filter_layout.addStretch()
        layout.addWidget(filter_bar)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setDocumentMode(True)

        self.trial_tab = QWidget()
        self.trial_layout = QVBoxLayout(self.trial_tab)
        self.trial_text = QTextEdit()
        self.trial_text.setReadOnly(True)
        self.trial_layout.addWidget(self.trial_text)
        self.tabs.addTab(self.trial_tab, "📊 ميزان المراجعة")

        self.income_tab = QWidget()
        self.income_layout = QVBoxLayout(self.income_tab)
        self.income_text = QTextEdit()
        self.income_text.setReadOnly(True)
        self.income_layout.addWidget(self.income_text)
        self.tabs.addTab(self.income_tab, "💰 قائمة الدخل")

        self.balance_tab = QWidget()
        self.balance_layout = QVBoxLayout(self.balance_tab)
        self.balance_text = QTextEdit()
        self.balance_text.setReadOnly(True)
        self.balance_layout.addWidget(self.balance_text)
        self.tabs.addTab(self.balance_tab, "🏦 الميزانية العمومية")

        self.customer_tab = QWidget()
        self.customer_layout = QVBoxLayout(self.customer_tab)
        customer_top = QHBoxLayout()
        self.customer_combo = QComboBox()
        self.customer_combo.setMinimumWidth(250)
        customer_top.addWidget(QLabel("اختر العميل:"))
        customer_top.addWidget(self.customer_combo)
        customer_top.addStretch()
        self.customer_layout.addLayout(customer_top)
        self.customer_text = QTextEdit()
        self.customer_text.setReadOnly(True)
        self.customer_layout.addWidget(self.customer_text)
        self.tabs.addTab(self.customer_tab, "👤 كشف حساب عميل")

        self.supplier_tab = QWidget()
        self.supplier_layout = QVBoxLayout(self.supplier_tab)
        supplier_top = QHBoxLayout()
        self.supplier_combo = QComboBox()
        self.supplier_combo.setMinimumWidth(250)
        supplier_top.addWidget(QLabel("اختر المورد:"))
        supplier_top.addWidget(self.supplier_combo)
        supplier_top.addStretch()
        self.supplier_layout.addLayout(supplier_top)
        self.supplier_text = QTextEdit()
        self.supplier_text.setReadOnly(True)
        self.supplier_layout.addWidget(self.supplier_text)
        self.tabs.addTab(self.supplier_tab, "🏭 كشف حساب مورد")

        layout.addWidget(self.tabs)
        self.refresh_customers()
        self.refresh_suppliers()
        self.refresh_all()

    def refresh_all(self):
        self.show_progress(True)
        self.show_trial_balance()
        self.show_income_statement()
        self.show_balance_sheet()
        self.refresh_customers()
        self.refresh_suppliers()
        self.show_progress(False)

    def show_progress(self, show):
        self.progress.setVisible(show)
        if show:
            self.progress.setRange(0, 0)
        QApplication.processEvents()

    def refresh_customers(self):
        customers = customer_dao.get_all()
        self.customer_combo.clear()
        for c in customers:
            self.customer_combo.addItem(f"{c.name} (الرصيد: {format_currency(c.balance)})", c.id)
        if customers:
            self.show_customer_statement()

    def refresh_suppliers(self):
        suppliers = supplier_dao.get_all()
        self.supplier_combo.clear()
        for s in suppliers:
            self.supplier_combo.addItem(f"{s.name} (الرصيد: {format_currency(s.balance)})", s.id)
        if suppliers:
            self.show_supplier_statement()

    def get_date_range(self):
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")
        return start, end

    def show_trial_balance(self):
        start, end = self.get_date_range()
        report = TrialBalanceReport().generate(start, end)
        self.trial_text.setHtml(report.to_html())

    def show_income_statement(self):
        start, end = self.get_date_range()
        report = IncomeStatementReport().generate(start, end)
        self.income_text.setHtml(report.to_html())

    def show_balance_sheet(self):
        start, end = self.get_date_range()
        report = BalanceSheetReport().generate(start, end)
        self.balance_text.setHtml(report.to_html())

    def show_customer_statement(self):
        cust_id = self.customer_combo.currentData()
        if not cust_id:
            self.customer_text.setHtml("<div style='text-align: center;'><h3>اختر عميلاً أولاً</h3></div>")
            return
        cust_name = self.customer_combo.currentText().split(' (')[0]
        report = CustomerStatementReport(cust_id, cust_name).generate()
        self.customer_text.setHtml(report.to_html())

    def show_supplier_statement(self):
        supp_id = self.supplier_combo.currentData()
        if not supp_id:
            self.supplier_text.setHtml("<div style='text-align: center;'><h3>اختر مورداً أولاً</h3></div>")
            return
        supp_name = self.supplier_combo.currentText().split(' (')[0]
        report = SupplierStatementReport(supp_id, supp_name).generate()
        self.supplier_text.setHtml(report.to_html())

    def get_current_html(self):
        current = self.tabs.currentWidget()
        if current == self.trial_tab:
            return self.trial_text.toHtml()
        elif current == self.income_tab:
            return self.income_text.toHtml()
        elif current == self.balance_tab:
            return self.balance_text.toHtml()
        elif current == self.customer_tab:
            return self.customer_text.toHtml()
        elif current == self.supplier_tab:
            return self.supplier_text.toHtml()
        return ""

    def export_current_to_pdf(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للتصدير", "error", self)
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير كـ PDF", "report.pdf", "PDF (*.pdf)")
        if not filename:
            return
        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(filename)
        doc.print(printer)
        show_toast("تم حفظ التقرير كـ PDF", "success", self)

    def export_current_to_excel(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للتصدير", "error", self)
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير كـ Excel", "report.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            show_centered_messagebox(self, "خطأ", "مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl", QMessageBox.Critical)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير"
        ws.cell(row=1, column=1, value="تم تصدير التقرير من نظام الراجحي")
        ws.cell(row=2, column=1, value="يرجى نسخ البيانات يدوياً من التقرير الأصلي")
        wb.save(filename)
        show_toast("تم حفظ التقرير كـ Excel (مبسط)", "success", self)

    def print_current_report(self):
        html = self.get_current_html()
        if not html:
            show_toast("لا يوجد تقرير للطباعة", "error", self)
            return
        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(lambda p: doc.print(p))
        preview.exec()
